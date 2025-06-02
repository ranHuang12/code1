# -*- coding: utf-8 -*-
"""
Created on Tue Aug 10 11:33:30 2021
LST质量筛选
F:\LST_006\China\MOD11A1\h28v05
@author: dell
"""

#encoding=utf-8
import datetime
start = datetime.datetime.now()

from pyhdf.SD import SD
import numpy as np
from scipy import optimize#提供许多优化数值算法
from scipy import linalg#线性代数操作，from scipy.linalg import solve求解线性方程组
from scipy.stats.mstats_basic import linregress

from osgeo import gdal
from osgeo.gdalconst import GDT_Float32

import os
import sys
import glob
import gc

import pandas as pd
import shutil
from osgeo import osr
from openpyxl import Workbook
from datetime import datetime
from sklearn.metrics import mean_squared_error
from sklearn import datasets,ensemble
import statsmodels.formula.api as smf


driver=gdal.GetDriverByName('GTiff')
driver.Register()

import calendar
def year_days(year):
    if calendar.isleap(year):
        day_all=366
    else:
        day_all=365
    return day_all
def CreateRaster(output_path,arr,proj,transform,nodata):
    h,w=np.shape(arr)
    oDS=driver.Create(output_path,w,h,1,GDT_Float32,['COMPRESS=DEFLATE'])
    outband1=oDS.GetRasterBand(1)
    for i in range(h):
        outband1.WriteArray(arr[i].reshape(1,-1),0,i)
    oDS.SetProjection(proj)
    oDS.SetGeoTransform(transform)
    outband1.FlushCache()
    outband1.SetNoDataValue(nodata)

def readRaster(file_path):
    #print (file_path)
    ds=gdal.Open(file_path)
    if ds==None:
        print ("failed to open file")
    band=ds.GetRasterBand(1)
    h=band.YSize; w=band.XSize
    proj=ds.GetProjection()
    transform=ds.GetGeoTransform()
    #arr=np.zeros((h,w))
    arr=band.ReadAsArray(0,0,w,h)
    return arr,proj,transform

def read_HDF_LST(HDF_file,select_type=1):
    """
    :param HDF_file HDF格式的MODIS LST数据
    :param select_type 质量筛选办法，等于或不等于1。等于1表示高质量筛选，不等于1表示有效值筛选
    return 返回经过质量筛选和单位转换后的夜间、白天LST
    """
     
    hdf = SD(HDF_file)
    #print(hdf.info())  # 信息类别数
    
    """data = hdf.datasets()
    for i in data:
        print(i)  # 具体类别
        img = hdf.select(i)[:]  # 图像数据,numpy数组
        plt.imshow(img, cmap='gray')  # 显示图像
        plt.show()"""
    arr_day_LST = hdf.select('LST_Day_1km')[:]#将LST数据读出为数组
    arr_night_LST = hdf.select('LST_Night_1km')[:]
    
    if select_type==1:#高质量筛选
        #将QC=0作为高质量筛选办法
        QC_day = hdf.select('QC_Day')[:]
        QC_day_select=np.zeros_like(QC_day)
        QC_day_select[QC_day==0]=1
        
        QC_night = hdf.select('QC_Night')[:]
        QC_night_select=np.zeros_like(QC_night)
        QC_night_select[QC_night==0]=1
        
        select_LST_day = QC_day_select*arr_day_LST*0.02-273.15#进行质量筛选，并将开式温度转为摄氏温度
        select_LST_day[select_LST_day==-273.15] = 0#将背景值还原为0
    
        select_LST_night = QC_night_select*arr_night_LST*0.02-273.15#进行质量筛选，并将开式温度转为摄氏温度
        select_LST_night[select_LST_night==-273.15] = 0#将背景值还原为0
    
    else: #有效值筛选，即不进行QC筛选
        select_LST_day = arr_day_LST*0.02-273.15
        select_LST_day[select_LST_day==-273.15] = 0
        
        select_LST_night = arr_night_LST*0.02-273.15
        select_LST_night[select_LST_night==-273.15] = 0
    return (select_LST_day,select_LST_night)

def read_HDF_LST_1(HDF_file,select_type=1):
    """
    :param HDF_file HDF格式的MODIS LST数据
    :param select_type 质量筛选办法，等于或不等于1。等于1表示高质量筛选，不等于1表示有效值筛选
    return 返回经过质量筛选和单位转换后的夜间、白天LST
    """
    path,hdf_file = os.path.split(HDF_file)
    name,ext=os.path.splitext(hdf_file)
    
    ds = gdal.Open(HDF_file)
    subdatasets = ds.GetSubDatasets()
    
    arr_day_LST = (gdal.Open(subdatasets[0][0])).ReadAsArray()
    QC_day = (gdal.Open(subdatasets[1][0])).ReadAsArray()
    
    arr_night_LST = (gdal.Open(subdatasets[4][0])).ReadAsArray()
    QC_night = (gdal.Open(subdatasets[5][0])).ReadAsArray()
    
    proj= gdal.Open(subdatasets[0][0]).GetProjection()
    transform= gdal.Open(subdatasets[0][0]).GetGeoTransform()
    
    LST_path = path + '\\LST_source\\'
    if not os.path.exists(LST_path):
        os.mkdir(LST_path)
    
    if select_type==1:#高质量筛选
        #将QC=0作为高质量筛选办法
        QC_day_select=np.zeros_like(QC_day)
        QC_day_select[QC_day==0]=1
        
        QC_night_select=np.zeros_like(QC_night)
        QC_night_select[QC_night==0]=1
        
        select_LST_day = QC_day_select*arr_day_LST*0.02-273.15#进行质量筛选，并将开式温度转为摄氏温度
        select_LST_day[select_LST_day==-273.15] = 0#将背景值还原为0
    
        select_LST_night = QC_night_select*arr_night_LST*0.02-273.15#进行质量筛选，并将开式温度转为摄氏温度
        select_LST_night[select_LST_night==-273.15] = 0#将背景值还原为0
    
    else: #有效值筛选，即不进行QC筛选
        select_LST_day = arr_day_LST*0.02-273.15
        select_LST_day[select_LST_day==-273.15] = 0
        
        select_LST_night = arr_night_LST*0.02-273.15
        select_LST_night[select_LST_night==-273.15] = 0
    
    name_list=name.split('.')
    list1 =[name_list[0][:3],name_list[1][1:],name_list[2]]
    name= ".".join(list1)#MOD.2013001.h26v05
    CreateRaster(LST_path+name+".day.tif",select_LST_day,proj,transform,0)
    CreateRaster(LST_path+name+".night.tif",select_LST_night,proj,transform,0)
        

def index_calculte(reflectance_file):
    """
            算相关指数，并进行重采样
    :param1 reflectance_file 反射率数据
    :param2 index_path 各种指数存放路径      
    """
    path,hdf_file = os.path.split(reflectance_file)
    name,ext=os.path.splitext(hdf_file)
    name_list=name.split('.')
    list1 =[name_list[0][0:3],name_list[1][1:],name_list[2]]
    name= ".".join(list1)
    
    ds = gdal.Open(reflectance_file)
    if ds==None:
        return
    subdatasets = ds.GetSubDatasets()
    red_ds = gdal.Open(subdatasets[7][0])
    nir1_ds = gdal.Open(subdatasets[8][0])
    blue_ds = gdal.Open(subdatasets[9][0])
    #green_ds = gdal.Open(subdatasets[10][0])
    #nir2_ds = gdal.Open(subdatasets[11][0])
    swir1_ds = gdal.Open(subdatasets[12][0])
    swir2_ds = gdal.Open(subdatasets[13][0])
    
    proj= red_ds.GetProjection()
    transform= red_ds.GetGeoTransform()
    
    EVI_path = path+"\\EVI\\"
    if not os.path.exists(EVI_path):
        os.mkdir(EVI_path)
    temp_path =  path+"\\temp\\"
    if not os.path.exists(temp_path):
        os.mkdir(temp_path)
    red_arr = 0.0001*(red_ds.ReadAsArray())
    nir1_arr =  0.0001*(nir1_ds.ReadAsArray())
    blue_arr = 0.0001*(blue_ds.ReadAsArray())
    cond = (red_arr==32767*0.0001)|(nir1_arr==32767*0.0001)|(blue_arr==32767*0.0001)
    EVI_arr = 2.5*((nir1_arr-red_arr)/(nir1_arr+6*red_arr-7.5*blue_arr+1))
    EVI_arr[cond] = -28672
    EVI_arr[np.where((EVI_arr<-1)|(EVI_arr>1))]=-28672
    CreateRaster(temp_path+name+".tif",EVI_arr,proj,transform,-28672)
    ds_evi=gdal.Open(temp_path+name+".tif")
    gdal.Warp(EVI_path+name+".tif",ds_evi,xRes=926.6254331, yRes=926.6254331)#重采样
    #evi_arr_1 = readRaster(EVI_path+name+".tif")[0]
    ds_evi = None
    os.remove(temp_path+name+".tif")
    
    NDFI_path = path+"\\NDFI\\"
    if not os.path.exists(NDFI_path):
        os.mkdir(NDFI_path)
    swir2_arr = 0.0001*(swir2_ds.ReadAsArray())
    cond = (red_arr==32767*0.0001)|(swir2_arr==32767*0.0001)
    NDFI_arr = (red_arr - swir2_arr)/(red_arr + swir2_arr)
    NDFI_arr[cond] = -28672
    CreateRaster(temp_path+name+".tif",NDFI_arr,proj,transform,-28672)
    ds_ndfi=gdal.Open(temp_path+name+".tif")
    gdal.Warp(NDFI_path+name+".tif",ds_ndfi,xRes=926.6254331, yRes=926.6254331)#重采样
    #ndfi_arr_1 = readRaster(NDFI_path+name+".tif")[0]
    ds_ndfi=None
    os.remove(temp_path+name+".tif")#删除中间文件
    
    NDBI_path = path+"\\NDBI\\"
    if not os.path.exists(NDBI_path):
        os.mkdir(NDBI_path)
    swir1_arr = 0.0001*(swir1_ds.ReadAsArray())
    cond = (nir1_arr==32767*0.0001)|(swir1_arr==32767*0.0001)
    NDBI_arr = (nir1_arr - swir1_arr)/(nir1_arr + swir1_arr)
    NDBI_arr[cond] = -28672
    CreateRaster(temp_path+name+".tif",NDBI_arr,proj,transform,-28672)
    ds_ndbi=gdal.Open(temp_path+name+".tif")
    gdal.Warp(NDBI_path+name+".tif",ds_ndbi,xRes=926.6254331, yRes=926.6254331)#重采样
    #ndbi_arr_1 = readRaster(NDBI_path+name+".tif")[0]
    ds_ndbi = None
    os.remove(temp_path+name+".tif")#删除中间文件
    #return evi_arr_1,ndbi_arr_1,ndbi_arr_1

def mosaic(data_list, out_path,nodata):

    # 读取其中一个栅格数据来确定镶嵌图像的一些属性
    if len(data_list)==0:
        return
    o_ds = gdal.Open(data_list[0])
    # 投影
    Projection = o_ds.GetProjection()
    # 波段数据类型
    o_ds_array = o_ds.ReadAsArray()

    if 'int8' in o_ds_array.dtype.name:
        datatype = gdal.GDT_Byte
    elif 'int16' in o_ds_array.dtype.name:
        datatype = gdal.GDT_UInt16
    else:
        datatype = gdal.GDT_Float32

    # 像元大小
    transform = o_ds.GetGeoTransform()
    pixelWidth = transform[1]
    pixelHeight = transform[5]

    del o_ds

    minx_list = []
    maxX_list = []
    minY_list = []
    maxY_list = []

    # 对于每一个需要镶嵌的数据，读取它的角点坐标
    for data in data_list:

        # 读取数据
        ds = gdal.Open(data)
        rows = ds.RasterYSize
        cols = ds.RasterXSize

        # 获取角点坐标
        transform = ds.GetGeoTransform()
        minX = transform[0]
        maxY = transform[3]
        pixelWidth = transform[1]
        pixelHeight = transform[5]  # 注意pixelHeight是负值
        maxX = minX + (cols * pixelWidth)
        minY = maxY + (rows * pixelHeight)

        minx_list.append(minX)
        maxX_list.append(maxX)
        minY_list.append(minY)
        maxY_list.append(maxY)

        del ds

    # 获取输出图像坐标
    minX = min(minx_list)
    maxX = max(maxX_list)
    minY = min(minY_list)
    maxY = max(maxY_list)

    # 获取输出图像的行与列
    cols = int(round((maxX - minX) / pixelWidth))
    rows = int(round((maxY - minY) / abs(pixelHeight)))# 注意pixelHeight是负值

    # 计算每个图像的偏移值
    xOffset_list = []
    yOffset_list = []
    i = 0

    for data in data_list:
        xOffset = int((minx_list[i] - minX) / pixelWidth)
        yOffset = int((maxY_list[i] - maxY) / pixelHeight)
        xOffset_list.append(xOffset)
        yOffset_list.append(yOffset)
        i += 1

    # 创建一个输出图像
    driver = gdal.GetDriverByName("GTiff")
    dsOut = driver.Create(out_path, cols, rows, 1, datatype) 
    bandOut = dsOut.GetRasterBand(1)
    bandOut.SetNoDataValue(nodata)

    i = 0
    #将原始图像写入新创建的图像
    for data in data_list:
        # 读取数据
        ds = gdal.Open(data)
        data_band = ds.GetRasterBand(1)
        data_rows = ds.RasterYSize
        data_cols = ds.RasterXSize

        data = data_band.ReadAsArray(0, 0, data_cols, data_rows)
        bandOut.WriteArray(data, xOffset_list[i], yOffset_list[i])
        

        del ds
        i += 1

    # 设置输出图像的几何信息和投影信息
    geotransform = [minX, pixelWidth, 0, maxY, 0, pixelHeight]
    dsOut.SetGeoTransform(geotransform)
    dsOut.SetProjection(Projection)
    bandOut.SetNoDataValue(0)

    del dsOut
def clip(tif_file,shpfile,outpath,nodata):
    src_ds=gdal.Open(tif_file)
    gdal.Warp(outpath,src_ds,cutlineDSName=shpfile,srcNodata=nodata, dstNodata=nodata)
    
def value_extract(sation_file,DEM_file,evi_path,ndfi_path,ndbi_path,lst_path,year,time,l_type):
    
    """
    :param1 sation_file: 站点数据，包括stationID,经纬度、X,Y
    :param2 dem_file:数据,如“F:\\case_study\\clip\\DEM\\mosaic_DEM_sin_1km.tif”
    :param3 evi_path:evi数据所在路径。如："F:\\case_study\\EVI\\"
    :param4 ndfi_path：NDFI数据所在路径。如：“F:\\case_study\\NDFI\\”
    :param5 ndbi_path:NDBI数据所在文件夹。如“F:\\case_study\\NDBI\\”
    :param6 lst_path:LST数据存放路径,如“F:\\case_study\\LST\\”
    :param7 year:年份
    :param8 time:LST成像时间，day or night
    :param9 l_type:LST类型，MOD或者MYD
    """ 

    arr_DEM,proj,transform = readRaster(DEM_file)
    
    df_OD=pd.DataFrame(columns=['StationID','Lon','Lat','X','Y',"DEM","EVI","NDFI","NDBI",'DOY','LST'])#建立一个空的dataframe作为表头
    data_path=sation_file
    if not os.path.exists("mete_data\\"):
        os.makedirs("mete_data\\")
    outpath="mete_data\\"
    for LST_file in os.listdir(lst_path):#MOD.2015006.day.tif
        name_list=LST_file.split('.')
        lst_type = name_list[0]
        lst_year = int((name_list[1])[0:4])
        date = name_list[1]
        DOY=int(date[-3:])
        lst_time = name_list[2]
        ext = name_list[-1]
        if ext!="tif" or lst_type!= l_type or lst_year!=year or lst_time!=time:
            continue 
        data_OD=pd.read_excel(data_path).iloc[:,0:]#将站点坐标读入pandas中
        data_OD["DEM"]=32767
        data_OD["EVI"] = -28672
        data_OD["NDFI"] = -28672
        data_OD["NDBI"] = -28672
        data_OD["DOY"]=-999
        data_OD["LST"]=-999
            
        evi_file_all = glob.glob(evi_path+"MCD."+str(date)+"*")
        if len(evi_file_all) == 0:
            continue
        evi_file = evi_file_all[0]
        evi_arr,proj,transform=readRaster(evi_file)
           
        ndfi_file_all = glob.glob(ndfi_path+"MCD."+str(date)+"*")
        if len(ndfi_file_all) == 0:
            continue
        ndfi_file = ndfi_file_all[0]
        ndfi_arr,proj,transform=readRaster(ndfi_file)
        
        ndbi_file_all = glob.glob(ndbi_path+"MCD."+str(date)+"*")
        if len(ndbi_file_all) == 0:
            continue
        ndbi_file = ndbi_file_all[0]
        ndbi_arr,proj,transform=readRaster(ndbi_file)
         
            
        #读入LST影像
        ds=gdal.Open(lst_path+LST_file)
        if ds == None:
            print (LST_file,"failed to open!")
            return
        width=ds.RasterXSize
        height=ds.RasterYSize
        Geotrans=ds.GetGeoTransform()
        Proj = ds.GetProjection() #获取投影信息
        x_pixel=Geotrans[1]
        y_pixel=Geotrans[5]
        X=[Geotrans[0],Geotrans[0]+width*x_pixel]#LST的X范围
        Y=[Geotrans[3]+height*y_pixel,Geotrans[3]]#LST的Y范围
        arr=ds.ReadAsArray(0,0,width,height)
        #获得LST影像各像元中心点坐标（类似于获得坐标点矢量）
        xgrid=np.arange(X[0]+x_pixel*0.5,X[1],x_pixel)
        ygrid=np.arange(Y[1]+y_pixel*0.5,Y[0],y_pixel)
    
        #根据excel表中站点的X,Y坐标来获得LST的值
        for i in range(0, len(data_OD)):#len(data)表示data的行数，注意表头作为dataframe不算入内
            if (xgrid[0] - x_pixel * 0.5) <= data_OD.iloc[i][3]  < (xgrid[-1] + x_pixel* 0.5) and (ygrid[-1] - x_pixel * 0.5)  <= data_OD.iloc[i][4] < (ygrid[0] + x_pixel * 0.5):
                data_OD.iloc[i,-2]=DOY
                if arr[int((ygrid[0] - data_OD.iloc[i][4]) / x_pixel + 0.5)][int((data_OD.iloc[i][3] - xgrid[0]) / x_pixel + 0.5)] != 0:
                    data_OD.iloc[i,-6]= arr_DEM[int((ygrid[0]-data_OD.iloc[i][4])/x_pixel+0.5)][int((data_OD.iloc[i][3] - xgrid[0]) / x_pixel + 0.5)]
                    data_OD.iloc[i,-5]= evi_arr[int((ygrid[0]-data_OD.iloc[i][4])/x_pixel+0.5)][int((data_OD.iloc[i][3] - xgrid[0]) / x_pixel + 0.5)]
                    data_OD.iloc[i,-4]= ndfi_arr[int((ygrid[0]-data_OD.iloc[i][4])/x_pixel+0.5)][int((data_OD.iloc[i][3] - xgrid[0]) / x_pixel + 0.5)]
                    data_OD.iloc[i,-3]= ndbi_arr[int((ygrid[0]-data_OD.iloc[i][4])/x_pixel+0.5)][int((data_OD.iloc[i][3] - xgrid[0]) / x_pixel + 0.5)]
                    data_OD.iloc[i,-1]= arr[int((ygrid[0]-data_OD.iloc[i][4])/x_pixel+0.5)][int((data_OD.iloc[i][3] - xgrid[0]) / x_pixel + 0.5)]#根据坐标获得相应值
                else:
                    data_OD.iloc[i,-6]= arr_DEM[int((ygrid[0]-data_OD.iloc[i][4])/x_pixel+0.5)][int((data_OD.iloc[i][3] - xgrid[0]) / x_pixel + 0.5)]
                    data_OD.iloc[i,-5]= evi_arr[int((ygrid[0]-data_OD.iloc[i][4])/x_pixel+0.5)][int((data_OD.iloc[i][3] - xgrid[0]) / x_pixel + 0.5)]
                    data_OD.iloc[i,-4]= ndfi_arr[int((ygrid[0]-data_OD.iloc[i][4])/x_pixel+0.5)][int((data_OD.iloc[i][3] - xgrid[0]) / x_pixel + 0.5)]
                    data_OD.iloc[i,-3]= ndbi_arr[int((ygrid[0]-data_OD.iloc[i][4])/x_pixel+0.5)][int((data_OD.iloc[i][3] - xgrid[0]) / x_pixel + 0.5)]
                    data_OD.iloc[i,-1]= -999#根据坐标获得相应值
                
        df_OD=pd.concat([df_OD,data_OD],ignore_index=True)
        print ("%d year %d day is done!" %(year,DOY))
        del data_OD  
    df1_OD=df_OD[~df_OD['LST'].isin([-999])]
    df1_OD.to_excel(outpath+str(year)+"_"+l_type+"_"+time+".xlsx",index=False)              
    gc.collect()

from openpyxl import Workbook
from openpyxl import load_workbook   
def LST_temp_merge(year_start,year_end,temp_path,l_type,time):
    
    """
    :param1 year_start: 建模数据开始年份
    :param2 year_end:建模数据结束年份
    :param3 f_temp:气象数据,包含station_ID，经纬度、X，Y。year,month,day,DOY,TA(日平均气温)
    :param4 l_type：LST数据类型
    :param5 time:LST数据时间
    """ 
    
    wb_spring=Workbook()
    ws_spring=wb_spring.active
    ws_spring.title=l_type+"_spring_LST_Temp"
    booktitle=["StationID","Lon","Lat","X","Y","year","month","day","DOY","DEM","EVI","NDFI","NDBI","TA","LST_Value"]
    ws_spring.append(booktitle)
    
    for year in range (year_start,year_end+1):
        f_LST="mete_data\\"+str(year)+"_"+l_type+"_"+time+".xlsx"
        wb_LST=load_workbook(f_LST)
        f_temp=temp_path+str(year)+"_TA_TL_TH.xlsx"
        wb_temp=load_workbook(f_temp)

        ws_LST=wb_LST.active
        ws_temp=wb_temp.active

        row_LST=ws_LST.max_row
        row_temp=ws_temp.max_row
        dictionry={}
        for i in range(2,row_LST+1):
            ID_LST=int(ws_LST.cell(row=i,column=1).value)
            DOY_LST=ws_LST.cell(row=i,column=10).value
            Lon=ws_LST.cell(row=i,column=2).value
            Lat=ws_LST.cell(row=i,column=3).value
            X=ws_LST.cell(row=i,column=4).value
            Y=ws_LST.cell(row=i,column=5).value
            dem = ws_LST.cell(row=i,column=6).value
            EVI =  ws_LST.cell(row=i,column=7).value
            NDFI =  ws_LST.cell(row=i,column=8).value
            NDBI = ws_LST.cell(row=i,column=9).value
            if EVI == -28672 or NDBI == -28672 or NDFI == -28672 or dem==32767:
                continue
            print ("LST_temp_merge")
            LST_value=ws_LST.cell(row=i,column=11).value
            key=str(ID_LST)+str(DOY_LST)
            row_value=[Lon,Lat,X,Y,dem,EVI,NDFI,NDBI,LST_value]
            dictionry[key]=row_value
    #气象站数据循环
        for j in range(2,row_temp+1):
            #气象数据站点ID
            ID_temp=ws_temp.cell(row=j,column=1).value
            year=int(ws_temp.cell(row=j,column=2).value)
            month=int(ws_temp.cell(row=j,column=3).value)
            day=int(ws_temp.cell(row=j,column=4).value)
            dt=datetime(year,month,day)
            DOY_temp=int(dt.strftime("%j"))
            TA=float(ws_temp.cell(row=j,column=5).value)
            the_key=str(ID_temp)+str(DOY_temp)
        
            if the_key not in dictionry:
                continue
            Lon_LST = dictionry[the_key][0]
            Lat_LST = dictionry[the_key][1]
            X_LST= dictionry[the_key][2]
            Y_LST= dictionry[the_key][3]
            DEM = dictionry[the_key][4]
            evi = dictionry[the_key][5]
            ndfi = dictionry[the_key][6]
            ndbi = dictionry[the_key][7]
            lst= dictionry[the_key][8]
            spring_value=[ID_temp,Lon_LST,Lat_LST,X_LST, Y_LST,year,month,day,DOY_temp,DEM,evi,ndfi,ndbi,TA,lst]
            ws_spring.append(spring_value)
            
    wb_spring.save("mete_data\\"+l_type+"_"+time+"_LST_temp.xlsx")


def rmse(y_test, y_true):
    return np.lib.scimath.sqrt(np.mean((y_test - y_true) ** 2))
def bias(y,y_simu):
    bias_sum1=[]
    for i in range(len(y)):
        bias_sum1.append(y[i]-y_simu[i])
    bias_value=np.mean(bias_sum1)
    return bias_value
def r2(ta,ta_simu):
    mean_ta=np.mean(ta)
    n=[]
    m=[]
    for i in range(len(ta_simu)):
        n.append(np.power(ta_simu[i]-mean_ta,2))
    for j in range(len(ta)):
        m.append(np.power(ta[j]-mean_ta,2))
    r2=float(np.sum(n))/np.sum(m)
    return r2

def muti_linear(year_start,year_end,model_data,dem_file,evi_path,ndfi_path,ndbi_path,lst_path,l_type,time,out_path):
    
    """
    :param1 year_start: 开始反演年份
    :param2 year_end:结束气温年份
    :param3 f_temp:气象数据,包含station_ID，经纬度、X，Y。year,month,day,DOY,TA(日平均气温)
    :param4 l_type：LST数据类型
    :param5 time:LST数据时间
    """
    df = pd.read_excel(model_data)
    sales_normal=df.iloc[:,[9,10,11,12,13,14]]#自变量
    y = df['TA'].values#因变量
    fit = smf.ols('TA~DEM+EVI+NDFI+NDBI+LST_Value', data = sales_normal).fit()#["StationID","Lon","Lat","X","Y","year","month","day","DOY","DEM","EVI","NDFI","NDBI","TA","LST_Value"]
    coef_Intercept= (fit.params).Intercept#截距
    coef_dem= (fit.params).DEM#DEM因子的系数
    coef_evi= (fit.params).EVI#EVI的系数
    coef_ndfi= (fit.params).EVI#NDFI的系数
    coef_ndbi= (fit.params).EVI#NDBI的系数
    coef_lst= (fit.params).LST_Value#LST的系数
    #以下为模型精度
    Y_pred1 = (fit.predict(exog = sales_normal)).values
    RMSE = rmse(Y_pred1, y)
    Bias = bias(y,Y_pred1)
    R2 = r2(y,Y_pred1)
    n=len(df)
    
    print ("建模精度：R2=%f,RMSE=%f,Bias=%f,样点总数=%d"%(R2,RMSE,Bias,n))

    dem_arr_n,proj,transform = readRaster(dem_file)
    
    for year in range(year_start,year_end+1):
        for lst_file in os.listdir(lst_path):
            name_list=lst_file.split('.')
            lst_type = name_list[0]
            lst_time= name_list[2]
            lst_year=int((name_list[1])[0:4])
            ext=name_list[-1]
            if ext !='tif' or lst_type!= l_type or time != lst_time or lst_year!=year:
                continue
            date=name_list[1]
            print (date)
            
            if os.path.exists(evi_path+"MCD."+date+".tif"):
                evi_arr_n,proj,transform=readRaster(evi_path+"MCD."+date+".tif")
            else:
                evi_arr_n=np.zeros_like(dem_arr_n)+-28672
            if os.path.exists(ndfi_path+"MCD."+date+".tif"):
                ndfi_arr_n,proj,transform=readRaster(ndfi_path+"MCD."+date+".tif")
            else:
                ndfi_arr_n=np.zeros_like(dem_arr_n)+-28672
            if os.path.exists(ndbi_path+"MCD."+date+".tif"):
                ndbi_arr_n,proj,transform=readRaster(ndbi_path+"MCD."+date+".tif")
            else:
                ndbi_arr_n=np.zeros_like(dem_arr_n)+-28672
                
            ds=gdal.Open(lst_path+lst_file)
            if ds == None:
                print (lst_file,"failed to open!")
                break
            width=ds.RasterXSize
            height=ds.RasterYSize
                        
            arr=ds.ReadAsArray(0,0,width,height)
            arr_TA=np.zeros_like(arr)
            outpath1 = out_path + l_type +"_LST_to_TA\\" 
            if not os.path.exists(outpath1):
                os.makedirs(outpath1)   
            out_name1=lst_file
            
            condi=(arr!=0)&(evi_arr_n!=-28672)&(ndfi_arr_n!=-28672)&(ndbi_arr_n!=-28672)&(dem_arr_n!=32767)
            arr_TA[condi]=coef_dem*dem_arr_n[condi]+coef_evi*evi_arr_n[condi]+coef_ndfi*ndfi_arr_n[condi]+coef_ndbi*ndbi_arr_n[condi]+coef_lst*arr[condi]+coef_Intercept#气温估算
            CreateRaster(outpath1+out_name1, arr_TA, proj, transform,0) 
                
    
def RF(year_start,year_end,model_data,dem_file,evi_path,ndfi_path,ndbi_path,lst_path,l_type,time,out_path):
    df = pd.read_excel(model_data)
    x = df.iloc[:,[9,10,11,12,14]].values
    y = df['TA'].values
    model_RandomForestRegressor_OD_spring = ensemble.RandomForestRegressor(n_estimators=20)#这里使用20个决策树
    model_RandomForestRegressor_OD_spring.fit(x, y)
    Y_pred1 = model_RandomForestRegressor_OD_spring.predict(x)
    RMSE = rmse(Y_pred1, y)
    Bias = bias(y,Y_pred1)
    R2 = r2(y,Y_pred1)
    n=len(df)
    
    print ("建模精度：R2=%f,RMSE=%f,Bias=%f,样点总数=%d"%(R2,RMSE,Bias,n))
    
    dem_ds=gdal.Open(dem_file)
    if dem_ds == None:
        print (dem_file,"failed to open!")
    width=dem_ds.RasterXSize
    height=dem_ds.RasterYSize
    DEM_arr=dem_ds.ReadAsArray(0,0,width,height)
    dem_arr_n=DEM_arr.reshape(-1,1)#转为一列
    
    for year in range(year_start,year_end+1):
        for lst_file in os.listdir(lst_path):
            name_list=lst_file.split('.')
            lst_type = name_list[0]
            lst_time= name_list[2]
            lst_year=int((name_list[1])[0:4])
            ext=name_list[-1]
            if ext !='tif' or lst_type!= l_type or time != lst_time or lst_year!=year:
                continue
            date=name_list[1]
            print (date)
            
            if not os.path.exists(evi_path+"MCD."+date+".tif"):
                evi_arr=np.zeros_like(DEM_arr)+-28672  
            else:
                ndfi_arr,proj,transform=readRaster(ndfi_path+"MCD."+date+".tif")
            evi_arr[evi_arr==-28672]=0
            arr_shape = evi_arr.shape
            evi_arr_n = evi_arr.reshape(-1,1)
            
            if not os.path.exists(ndfi_path+"MCD."+date+".tif"):
                ndfi_arr=np.zeros_like(DEM_arr)+-28672  
            else:
                ndfi_arr,proj,transform=readRaster(ndfi_path+"MCD."+date+".tif")
            ndfi_arr[ndfi_arr==-28672]=0
            arr_shape = ndfi_arr.shape
            ndfi_arr_n = ndfi_arr.reshape(-1,1)
            
    
            if not os.path.exists(ndbi_path+"MCD."+date+".tif"):
                ndfi_arr=np.zeros_like(DEM_arr)+-28672  
            else:
                ndbi_arr,proj,transform=readRaster(ndbi_path+"MCD."+date+".tif")
            ndbi_arr[ndfi_arr==-28672]=0
            arr_shape = ndbi_arr.shape
            ndbi_arr_n = ndbi_arr.reshape(-1,1)
                
            ds=gdal.Open(lst_path+lst_file)
            if ds == None:
                print (lst_file,"failed to open!")
                break
            width=ds.RasterXSize
            height=ds.RasterYSize
                        
            lst_arr=ds.ReadAsArray(0,0,width,height)
            arr = lst_arr.reshape(-1,1)
            
            TA_data = np.hstack((dem_arr_n,evi_arr_n,ndfi_arr_n,ndbi_arr_n,arr))
            TA =model_RandomForestRegressor_OD_spring.predict(TA_data)
            arr_TA = TA.reshape(arr_shape)
            a=np.ones_like(arr_TA)
            condi=(arr!=0)&(evi_arr_n!=-28672)&(ndfi_arr_n!=-28672)&(ndbi_arr_n!=-28672)&(dem_arr_n!=32767)
            a[condi]=0
            arr_TA*=a
            
            outpath1 = out_path + l_type +"_LST_to_TA\\" 
            if not os.path.exists(outpath1):
                os.makedirs(outpath1)   
            out_name1=lst_file
            CreateRaster(outpath1+out_name1, arr_TA, proj, transform,0)  
    
def main(argv):
    import argparse 
    DESC = "Estimation of temperature by LST data"   
    parser = argparse.ArgumentParser(prog=argv[0], description=DESC)
    
    parser.add_argument('-start', '--start', dest='start', metavar='Number', help='start year of modeling data', required=True)
    parser.add_argument('-end', '--end', dest='end', metavar='Number', help='end year of modeling data', required=True)
    parser.add_argument('-year', '--year', dest='year', metavar='Number list', help='years for temperature estimation', required=True)#[2005,2009]表示分别对2005、2006、2007、2008年分数据进行气温估算
    parser.add_argument('-lst', '--lst', dest='lst', metavar='DIR', help='Input path of lst data',required=True)
    parser.add_argument('-reflec', '--reflec', dest='reflec', metavar='DIR', help='Input path of reflectance data for EVI/NDFI/NDBI',required=True)
    parser.add_argument('-dem', '--dem', dest='dem', metavar='DIR', help='Input path of dem data', required=True)
    parser.add_argument('-shp', '--shpfile', dest='shp', metavar='DIR', help='Input path of vector boundary file', required=True)
    parser.add_argument('-station', '--station', dest='station', metavar='file', help='Input path of meteorological station file', required=True)
    parser.add_argument('-mete', '--mete', dest='mete', metavar='DIR', help='Input path of meteorological data', required=True)
    parser.add_argument('-out', '--out', dest='out', metavar='DIR', help='Storage path of temperature data after estimation', required=True)
    parser.add_argument('-tile', '--tile', dest='tile', metavar='list',default=['h28v06','h28v05'], help="Tile list of LST data,default is ['h28v06','h28v05']")
    parser.add_argument('-l_type', '--lst_type', dest='lst_type', metavar='list',default=['MOD','MYD'], help="Lst data type list, 'MOD' for Terra satellite,'MYD' for Aqua satellite,default is ['MOD','MYD']")
    parser.add_argument('-time', '--lst_time', dest='lst_time', metavar='Number list',default=[0,1], help="Lst imaging time list, 0 for day, 1 for night, the default is [0,1]")
    parser.add_argument('-s_type', '--select_type', dest='select_type', metavar='Number',default='1', help="Lst quality screening method, 1 for high quality , not 1 for all valid value. The default is 1")
    parser.add_argument('-model', '--model_type', dest='model', metavar='Name',default='1', help="model for temperature estimation,1 for multi-linear model,2 for random forest.The default is 1")
    
    args = parser.parse_args(argv[1:])
    year_start = eval(args.start)#开始年份
    year_end = eval(args.end)
    
    year_list = eval(args.year)
    
    lst_path = args.lst
    if not os.path.exists(lst_path):
        print("输入的LST数据不存在，请检查！！！")
        os._exit(0)
            
            
    out_path = args.out
    if not os.path.exists(out_path):
        os.makedirs(out_path)        
           
    tile_list = eval(args.tile)
    lst_type_list = eval(args.lst_type)
    lst_time_list = eval(args.lst_time)
    select_type = eval(args.select_type)
    model_type = eval(args.model) 
    
    print ("开始预处理...")
    print ('LST数据开始进行质量筛选')
    
    for hdf in os.listdir(lst_path):
        name,ext=os.path.splitext(hdf)
        if ext!='.hdf':
           continue 
        HDF_file = lst_path +  hdf
        read_HDF_LST_1(HDF_file,select_type)#将经过质量筛选的LST数据存在lst_path + '\\LST_source\\'路径下
  
if __name__ == '__main__':
    try:
        sys.exit(main(sys.argv))
    except KeyboardInterrupt:
        sys.exit(-1)
    sys.exit(0)
     
                
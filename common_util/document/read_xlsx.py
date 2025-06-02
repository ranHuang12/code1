import openpyxl
import csv
import openpyxl
import xlrd
import numpy as np

xlsx1 = openpyxl.load_workbook(r'C:\Users\dell\Downloads\maize_wheat_statistic_pixel_number.xlsx')
xlsx2 = openpyxl.load_workbook(r'E:\Project_of_Hebei\20211018河北省各县编号.xlsx')
sheet1 = xlsx1['wheat']
sheet2 = xlsx2['Sheet1']
# for i in range(2, 233):
#     tile = sheet.cell(i, 2).value
#     location = sheet.cell(i, 3).value
#     if location != '无':
#         f.write(tile+'\n')
#         AOI[tile] = location
# for year in range(2015, 2020):
#     data = []
#     for i in range(3, 172):
#         line = []
#         line.append(sheet1.cell(i, 1).value)
#         line.append(sheet2.cell(i - 1, 2).value)
#         line.append(sheet1.cell(i, year-2013).value*900/10000)
#         data.append(line)
#
#     with open(r'C:\Users\dell\Downloads\河北省%s年县级小麦播种面积统计表.csv' % year, 'w', newline='') as f:
#         line_writer = csv.writer(f)
#         line1 = ['Number', 'CountyName', 'Area(ha)']
#         line_writer.writerow(line1)
#         for line in data:
#             line_writer.writerow(line)

excel_filepath = r'C:\Users\dell\Downloads\maize_wheat_statistic_pixel_number.xlsx'
ExcelFile = xlrd.open_workbook(excel_filepath)
sheet = ExcelFile.sheet_by_name('wheat')
cols_reference = sheet.col_values(2018-2014)  # 定位至统计表Area列，注意从公顷单位换算为像元单位(30m为 Area*10000/900)
cols_reference1 = [float(row)*900/10000000 for row in cols_reference[2:]]
print(cols_reference1)

reader = csv.reader(open(r'E:\Project_of_Hebei\河北省2018年县级小麦播种面积统计表.csv', 'r'))
next(reader)
cols_reference = [float(row[2])/1000 for row in reader]
print(cols_reference)
